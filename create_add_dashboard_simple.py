import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

print("Creating simplified ADD dashboard...")

# Load and prepare data
df = pd.read_excel('Compiled IPD case data SIMSRH_4months.xls')
df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

# Extract admission date
def extract_date_from_ip(ip_num):
    if pd.isna(ip_num):
        return None
    ip_str = str(ip_num)
    if len(ip_str) >= 9 and ip_str.startswith('IP'):
        try:
            year = 2000 + int(ip_str[2:4])
            month = int(ip_str[4:6])
            day = int(ip_str[6:8])
            return datetime(year, month, day)
        except:
            return None
    return None

df['admission_date'] = df['ip_number'].apply(extract_date_from_ip)

# Parse demographics
df['age'] = df['a/s'].str.extract(r'(\d+)').astype(float)
df['gender'] = df['a/s'].str.extract(r'/([MF])')

# Find ADD cases
add_keywords = ['gastroenteritis', 'gastro', 'diarrhea', 'diarrhoea', 'diarrh', 'dysentery', 'cholera', 'food poisoning', 'add', 'acute ge', 'age', 'diarrhrea', 'loose', 'motion', 'stool', 'bowel', 'enteric', 'dehydration']
add_cases = []
for idx, row in df.iterrows():
    diagnosis = str(row['diagnosis']).lower()
    if any(keyword in diagnosis for keyword in add_keywords):
        add_cases.append(idx)

add_df = df.loc[add_cases].copy() if add_cases else pd.DataFrame()
print(f"Found {len(add_df)} ADD cases")

# Create age groups
if len(add_df) > 0:
    age_bins = [0, 5, 18, 35, 50, 65, 100]
    age_labels = ['0-4', '5-17', '18-34', '35-49', '50-64', '65+']
    add_df['age_group'] = pd.cut(add_df['age'], bins=age_bins, labels=age_labels, right=False)

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Sheet 1: Dashboard Summary
ws_summary = wb.create_sheet("Summary")

ws_summary['A1'] = "SIMSRH IPD - Acute Diarrheal Disease (ADD) Dashboard"
ws_summary['A1'].font = Font(size=16, bold=True)
ws_summary.merge_cells('A1:D1')

# Key metrics
ws_summary['A3'] = "Key Metrics"
ws_summary['A3'].font = Font(size=14, bold=True)

metrics = [
    ["Total ADD Cases", len(add_df)],
    ["Percentage of Admissions", f"{len(add_df)/len(df)*100:.1f}%"],
    ["Date Range", "Aug 1 - Nov 12, 2025"],
    ["Mean Age", f"{add_df['age'].mean():.1f}"],
    ["Median Age", f"{add_df['age'].median():.1f}"],
    ["Male Cases", (add_df['gender'] == 'M').sum()],
    ["Female Cases", (add_df['gender'] == 'F').sum()],
]

for i, (metric, value) in enumerate(metrics, 5):
    ws_summary[f'A{i}'] = metric
    ws_summary[f'B{i}'] = value
    ws_summary[f'A{i}'].font = Font(bold=True)

# Sheet 2: Raw Data
ws_data = wb.create_sheet("Raw Data")
data_cols = ['ip_number', 'diagnosis', 'department', 'age', 'gender', 'admission_date']
for r, row in enumerate(dataframe_to_rows(add_df[data_cols], index=False), 1):
    for c, value in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=value)

# Sheet 3: Statistics
ws_stats = wb.create_sheet("Statistics")

# Age statistics
ws_stats['A1'] = "Age Statistics"
ws_stats['A1'].font = Font(size=14, bold=True)

age_stats = [
    ["Mean Age", add_df['age'].mean()],
    ["Median Age", add_df['age'].median()],
    ["Std Deviation", add_df['age'].std()],
    ["Min Age", add_df['age'].min()],
    ["Max Age", add_df['age'].max()],
    ["Count", len(add_df['age'].dropna())],
]

for i, (stat, value) in enumerate(age_stats, 2):
    ws_stats[f'A{i}'] = stat
    ws_stats[f'B{i}'] = value

# Gender distribution
ws_stats['D1'] = "Gender Distribution"
ws_stats['D1'].font = Font(size=14, bold=True)

gender_stats = [
    ["Male", (add_df['gender'] == 'M').sum(), f"{(add_df['gender'] == 'M').sum()/len(add_df)*100:.1f}%"],
    ["Female", (add_df['gender'] == 'F').sum(), f"{(add_df['gender'] == 'F').sum()/len(add_df)*100:.1f}%"],
    ["Total", len(add_df), "100.0%"],
]

for i, (gender, count, pct) in enumerate(gender_stats, 2):
    ws_stats[f'D{i}'] = gender
    ws_stats[f'E{i}'] = count
    ws_stats[f'F{i}'] = pct

# Age group distribution
ws_stats['H1'] = "Age Group Distribution"
ws_stats['H1'].font = Font(size=14, bold=True)

if len(add_df) > 0:
    age_group_stats = add_df['age_group'].value_counts().sort_index()
    for i, (group, count) in enumerate(age_group_stats.items(), 2):
        ws_stats[f'H{i}'] = str(group)
        ws_stats[f'I{i}'] = count
        ws_stats[f'J{i}'] = f"{count/len(add_df)*100:.1f}%"

# Department distribution
ws_stats['A15'] = "Department Distribution"
ws_stats['A15'].font = Font(size=14, bold=True)

dept_stats = add_df['department'].value_counts()
for i, (dept, count) in enumerate(dept_stats.items(), 16):
    ws_stats[f'A{i}'] = dept
    ws_stats[f'B{i}'] = count
    ws_stats[f'C{i}'] = f"{count/len(add_df)*100:.1f}%"

# Top diagnoses
ws_stats['E15'] = "Top 10 Diagnoses"
ws_stats['E15'].font = Font(size=14, bold=True)

top_dx = add_df['diagnosis'].value_counts().head(10)
for i, (dx, count) in enumerate(top_dx.items(), 16):
    ws_stats[f'E{i}'] = dx[:50] + "..." if len(dx) > 50 else dx  # Truncate long diagnoses
    ws_stats[f'F{i}'] = count
    ws_stats[f'G{i}'] = f"{count/len(add_df)*100:.1f}%"

# Sheet 4: Charts Data
ws_charts = wb.create_sheet("Charts Data")

# Age group data
ws_charts['A1'] = "Age Group Distribution"
if len(add_df) > 0:
    age_group_data = add_df['age_group'].value_counts().sort_index().reset_index()
    age_group_data.columns = ['Age Group', 'Count']
    for r, row in enumerate(dataframe_to_rows(age_group_data, index=False), 2):
        for c, value in enumerate(row, 1):
            ws_charts.cell(row=r, column=c, value=value)

# Department data
ws_charts['A10'] = "Department Distribution"
dept_data = add_df['department'].value_counts().reset_index()
dept_data.columns = ['Department', 'Count']
for r, row in enumerate(dataframe_to_rows(dept_data, index=False), 11):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Save workbook
wb.save('ADD_Dashboard.xlsx')
print("ADD Dashboard Excel file created successfully!")

# Create charts
plt.style.use('default')
sns.set_palette("husl")

# Age distribution
plt.figure(figsize=(10, 6))
plt.hist(add_df['age'].dropna(), bins=15, edgecolor='black', alpha=0.7, color='skyblue')
plt.xlabel('Age (years)')
plt.ylabel('Number of ADD Cases')
plt.title('Age Distribution of Acute Diarrheal Disease Cases')
plt.grid(True, alpha=0.3)
plt.axvline(add_df['age'].mean(), color='red', linestyle='--', linewidth=2,
           label=f'Mean: {add_df["age"].mean():.1f} years')
plt.legend()
plt.tight_layout()
plt.savefig('add_age_distribution.png', dpi=300, bbox_inches='tight')
plt.close()

# Gender pie chart
plt.figure(figsize=(8, 8))
gender_counts = add_df['gender'].value_counts()
colors = ['lightblue', 'lightcoral']
explode = (0.05, 0)

plt.pie(gender_counts.values, labels=gender_counts.index, autopct='%1.1f%%',
        colors=colors, explode=explode, shadow=True, startangle=90)
plt.title('Gender Distribution in ADD Cases', fontsize=14, fontweight='bold')
plt.axis('equal')
plt.tight_layout()
plt.savefig('add_gender_pie.png', dpi=300, bbox_inches='tight')
plt.close()

# Age group bar chart
if len(add_df) > 0:
    plt.figure(figsize=(10, 6))
    age_group_counts = add_df['age_group'].value_counts().sort_index()
    bars = plt.bar(range(len(age_group_counts)), age_group_counts.values, color='orange', edgecolor='black', alpha=0.8)
    plt.xticks(range(len(age_group_counts)), age_group_counts.index, rotation=45)
    plt.xlabel('Age Group')
    plt.ylabel('Number of Cases')
    plt.title('ADD Cases by Age Group')
    plt.grid(True, alpha=0.3)

    for bar, count in zip(bars, age_group_counts.values):
        plt.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.5,
                 f'{int(count)}', ha='center', va='bottom', fontweight='bold')

    plt.tight_layout()
    plt.savefig('add_age_groups.png', dpi=300, bbox_inches='tight')
    plt.close()

print("ADD dashboard and charts created successfully!")
print("Files created:")
print("- ADD_Dashboard.xlsx")
print("- add_age_distribution.png")
print("- add_gender_pie.png")
print("- add_age_groups.png")

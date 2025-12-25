import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Load and prepare data
print("Loading data and creating ADD dashboard...")
df = pd.read_excel('Compiled IPD case data SIMSRH_4months.xls')
df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

# Extract correct admission date from IP Number
def extract_date_from_ip(ip_num):
    if pd.isna(ip_num):
        return None
    ip_str = str(ip_num)
    if len(ip_str) >= 9 and ip_str.startswith('IP'):
        try:
            year = 2000 + int(ip_str[2:4])
            month = int(ip_str[4:6])
            day = int(ip_str[6:8])
            return datetime(year, month, day)
        except:
            return None
    return None

df['admission_date'] = df['ip_number'].apply(extract_date_from_ip)
df['admission_time_only'] = pd.to_datetime(df['admission_time'], errors='coerce').dt.time
df['admission_datetime'] = pd.to_datetime(df.apply(
    lambda row: f"{row['admission_date'].strftime('%Y-%m-%d')} {row['admission_time_only']}" if pd.notna(row['admission_date']) and pd.notna(row['admission_time_only']) else None, axis=1
), errors='coerce')
df['discharge_time'] = pd.to_datetime(df['discharge_time'], errors='coerce')

# Parse demographics
df['age'] = df['a/s'].str.extract(r'(\d+)').astype(float)
df['gender'] = df['a/s'].str.extract(r'/([MF])')

# Calculate LOS
df['length_of_stay'] = (df['discharge_time'] - df['admission_datetime']).dt.total_seconds() / (24 * 3600)

# Comprehensive search for gastroenteritis cases (ADD)
gi_keywords = ['gastroenteritis', 'gastro', 'diarrhea', 'diarrhoea', 'diarrh', 'dysentery', 'cholera', 'food poisoning', 'add', 'acute ge', 'age', 'diarrhrea', 'loose', 'motion', 'stool', 'bowel', 'enteric', 'dehydration']
gi_cases = []
for idx, row in df.iterrows():
    diagnosis = str(row['diagnosis']).lower()
    if any(keyword in diagnosis for keyword in gi_keywords):
        gi_cases.append(idx)

gi_df = df.loc[gi_cases].copy() if gi_cases else pd.DataFrame()

# Create age groups
if len(gi_df) > 0:
    age_bins = [0, 5, 18, 35, 50, 65, 100]
    age_labels = ['0-4', '5-17', '18-34', '35-49', '50-64', '65+']
    gi_df['age_group'] = pd.cut(gi_df['age'], bins=age_bins, labels=age_labels, right=False)

print(f"Found {len(gi_df)} ADD cases for dashboard creation")

# Create Excel workbook
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Sheet 1: Dashboard Summary
ws_dashboard = wb.create_sheet("Dashboard Summary")

# Title
ws_dashboard['A1'] = "SIMSRH IPD - Acute Diarrheal Disease (ADD) Dashboard"
ws_dashboard['A1'].font = Font(size=16, bold=True)
ws_dashboard.merge_cells('A1:G1')

# Key Metrics
ws_dashboard['A3'] = "Key Metrics"
ws_dashboard['A3'].font = Font(size=14, bold=True)

headers = ['Metric', 'Value', 'Formula/Notes']
for col, header in enumerate(headers, 1):
    ws_dashboard.cell(row=4, column=col).value = header
    ws_dashboard.cell(row=4, column=col).font = Font(bold=True)
    ws_dashboard.cell(row=4, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Summary data
summary_data = [
    ["Total ADD Cases", len(gi_df), f"=COUNTA('Raw Data'!A:A)-1"],
    ["Percentage of Total Admissions", ".1f", f"=B5/COUNTA('All Admissions'!A:A)*100"],
    ["Date Range", "August 1 - November 12, 2025", ""],
    ["Mean Age (SD)", ".1f", "=AVERAGE('Raw Data'!F:F)"],
    ["Median Age", ".1f", "=MEDIAN('Raw Data'!F:F)"],
    ["Male Cases", (gi_df['gender'] == 'M').sum(), "=COUNTIF('Raw Data'!G:G,\"M\")"],
    ["Female Cases", (gi_df['gender'] == 'F').sum(), "=COUNTIF('Raw Data'!G:G,\"F\")"],
    ["Average Length of Stay", ".1f", "=AVERAGE('Raw Data'!H:H)"],
]

for row, data in enumerate(summary_data, 5):
    for col, value in enumerate(data, 1):
        if col == 2 and isinstance(value, str) and value.endswith('f'):
            # Handle formatted values
            if "Mean Age" in data[0]:
                ws_dashboard.cell(row=row, column=col).value = f"{gi_df['age'].mean():.1f} ({gi_df['age'].std():.1f})"
            elif "Median Age" in data[0]:
                ws_dashboard.cell(row=row, column=col).value = f"{gi_df['age'].median():.1f}"
            elif "Average Length" in data[0]:
                ws_dashboard.cell(row=row, column=col).value = f"{gi_df['length_of_stay'].mean():.1f}"
            elif "Percentage" in data[0]:
                ws_dashboard.cell(row=row, column=col).value = ".1f"
        else:
            ws_dashboard.cell(row=row, column=col).value = value

# Auto-adjust column widths
for col in ws_dashboard.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_dashboard.column_dimensions[column].width = adjusted_width

# Sheet 2: Raw Data
ws_raw = wb.create_sheet("Raw Data")

# Write GI dataframe to Excel
for r, row in enumerate(dataframe_to_rows(gi_df[['ip_number', 'diagnosis', 'department', 'age', 'gender', 'length_of_stay', 'admission_datetime']], index=False), 1):
    for c, value in enumerate(row, 1):
        ws_raw.cell(row=r, column=c, value=value)

# Sheet 3: Summary Statistics
ws_stats = wb.create_sheet("Summary Statistics")

# Age statistics
ws_stats['A1'] = "Age Statistics"
ws_stats['A1'].font = Font(size=14, bold=True)

age_stats_data = [
    ["Statistic", "Value"],
    ["Count", len(gi_df['age'].dropna())],
    ["Mean", gi_df['age'].mean()],
    ["Std Dev", gi_df['age'].std()],
    ["Min", gi_df['age'].min()],
    ["25%", gi_df['age'].quantile(0.25)],
    ["Median", gi_df['age'].median()],
    ["75%", gi_df['age'].quantile(0.75)],
    ["Max", gi_df['age'].max()],
]

for row, data in enumerate(age_stats_data, 1):
    for col, value in enumerate(data, 1):
        ws_stats.cell(row=row, column=col).value = value
        if row == 1:
            ws_stats.cell(row=row, column=col).font = Font(bold=True)

# Gender distribution
ws_stats['A12'] = "Gender Distribution"
ws_stats['A12'].font = Font(size=14, bold=True)

gender_data = [
    ["Gender", "Count", "Percentage"],
    ["Male", (gi_df['gender'] == 'M').sum(), ".1f"],
    ["Female", (gi_df['gender'] == 'F').sum(), ".1f"],
    ["Total", len(gi_df), "100.0"],
]

for row, data in enumerate(gender_data, 13):
    for col, value in enumerate(data, 1):
        if col == 3 and value == ".1f":
            total = len(gi_df)
            if "Male" in data[0]:
                ws_stats.cell(row=row, column=col).value = (gi_df['gender'] == 'M').sum() / total * 100
            elif "Female" in data[0]:
                ws_stats.cell(row=row, column=col).value = (gi_df['gender'] == 'F').sum() / total * 100
            else:
                ws_stats.cell(row=row, column=col).value = 100.0
        else:
            ws_stats.cell(row=row, column=col).value = value
        if row == 13:
            ws_stats.cell(row=row, column=col).font = Font(bold=True)

# Department distribution
ws_stats['A18'] = "Department Distribution"
ws_stats['A18'].font = Font(size=14, bold=True)

dept_dist = gi_df['department'].value_counts().reset_index()
dept_dist.columns = ['Department', 'Count']

dept_data = [["Department", "Count", "Percentage"]]
for _, row in dept_dist.iterrows():
    dept_data.append([row['Department'], row['Count'], ".1f"])

for row, data in enumerate(dept_data, 19):
    for col, value in enumerate(data, 1):
        if col == 3 and value == ".1f":
            ws_stats.cell(row=row, column=col).value = data[1] / len(gi_df) * 100
        else:
            ws_stats.cell(row=row, column=col).value = value
        if row == 19:
            ws_stats.cell(row=row, column=col).font = Font(bold=True)

# Sheet 4: Charts Data
ws_charts = wb.create_sheet("Charts Data")

# Age group distribution
ws_charts['A1'] = "Age Group Distribution"
ws_charts['A1'].font = Font(size=14, bold=True)

age_group_dist = gi_df['age_group'].value_counts().sort_index().reset_index()
age_group_dist.columns = ['Age Group', 'Count']

for r, row in enumerate(dataframe_to_rows(age_group_dist, index=False), 2):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Top diagnoses
ws_charts['A10'] = "Top 10 Diagnoses"
ws_charts['A10'].font = Font(size=14, bold=True)

top_diagnoses = gi_df['diagnosis'].value_counts().head(10).reset_index()
top_diagnoses.columns = ['Diagnosis', 'Count']

for r, row in enumerate(dataframe_to_rows(top_diagnoses, index=False), 11):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Create Charts sheet with embedded charts
ws_visualizations = wb.create_sheet("Visualizations")

# Create age distribution chart
chart1 = BarChart()
chart1.title = "Age Group Distribution - ADD Cases"
chart1.y_axis.title = "Number of Cases"
chart1.x_axis.title = "Age Group"

data = Reference(ws_charts, min_col=2, min_row=2, max_row=len(age_group_dist)+1)
cats = Reference(ws_charts, min_col=1, min_row=3, max_row=len(age_group_dist)+1)
chart1.add_data(data, titles_from_data=False)
chart1.set_categories(cats)
chart1.width = 15
chart1.height = 10

ws_visualizations.add_chart(chart1, "A1")

# Create diagnosis distribution chart
chart2 = BarChart()
chart2.title = "Top 10 ADD Diagnoses"
chart2.y_axis.title = "Number of Cases"
chart2.x_axis.title = "Diagnosis"

data2 = Reference(ws_charts, min_col=2, min_row=11, max_row=20)
cats2 = Reference(ws_charts, min_col=1, min_row=12, max_row=20)
chart2.add_data(data2, titles_from_data=False)
chart2.set_categories(cats2)
chart2.width = 20
chart2.height = 12

ws_visualizations.add_chart(chart2, "A20")

# Sheet 5: All Admissions Data (for reference)
ws_all = wb.create_sheet("All Admissions")

# Write sample of all admissions data
sample_df = df[['ip_number', 'diagnosis', 'department', 'age', 'gender']].head(100)  # First 100 rows
for r, row in enumerate(dataframe_to_rows(sample_df, index=False), 1):
    for c, value in enumerate(row, 1):
        ws_all.cell(row=r, column=c, value=value)

# Add note
ws_all['A102'] = "Note: This is a sample of the first 100 admissions from the full dataset for reference."

# Save the workbook
wb.save('ADD_Analysis_Dashboard.xlsx')
print("ADD dashboard Excel file created successfully: ADD_Analysis_Dashboard.xlsx")

# Create additional charts as images for embedding
plt.style.use('default')
sns.set_palette("husl")

# Age distribution histogram
plt.figure(figsize=(10, 6))
plt.hist(gi_df['age'].dropna(), bins=15, edgecolor='black', alpha=0.7, color='skyblue')
plt.xlabel('Age (years)')
plt.ylabel('Number of ADD Cases')
plt.title('Age Distribution of Acute Diarrheal Disease Cases at SIMSRH')
plt.grid(True, alpha=0.3)
plt.axvline(gi_df['age'].mean(), color='red', linestyle='--', linewidth=2,
           label=f'Mean: {gi_df["age"].mean():.1f} years')
plt.legend()
plt.tight_layout()
plt.savefig('add_age_distribution.png', dpi=300, bbox_inches='tight')
plt.close()

# Gender pie chart
plt.figure(figsize=(8, 8))
gender_counts = gi_df['gender'].value_counts()
colors = ['lightblue', 'lightcoral']
explode = (0.05, 0)

plt.pie(gender_counts.values, labels=gender_counts.index, autopct='%1.1f%%',
        colors=colors, explode=explode, shadow=True, startangle=90)
plt.title('Gender Distribution in Acute Diarrheal Disease Cases at SIMSRH', fontsize=14, fontweight='bold')
plt.axis('equal')
plt.tight_layout()
plt.savefig('add_gender_distribution.png', dpi=300, bbox_inches='tight')
plt.close()

print("ADD dashboard creation completed!")
print("Files created:")
print("- ADD_Analysis_Dashboard.xlsx (comprehensive Excel dashboard)")
print("- add_age_distribution.png (age distribution chart)")
print("- add_gender_distribution.png (gender distribution chart)")

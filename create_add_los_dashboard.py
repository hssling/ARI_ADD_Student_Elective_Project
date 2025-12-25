import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

print("Creating ADD-specific Length of Stay (LOS) dashboard...")

# Load and prepare data
df = pd.read_excel('Compiled IPD case data SIMSRH_4months.xls')
df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

# Extract admission date from IP Number
def extract_date_from_ip(ip_num):
    if pd.isna(ip_num):
        return None
    ip_str = str(ip_num)
    if len(ip_str) >= 9 and ip_str.startswith('IP'):
        try:
            year = 2000 + int(ip_str[2:4])
            month = int(ip_str[4:6])
            day = int(ip_str[6:8])
            return datetime(year, month, day)
        except:
            return None
    return None

df['admission_date'] = df['ip_number'].apply(extract_date_from_ip)
df['admission_time_only'] = pd.to_datetime(df['admission_time'], errors='coerce').dt.time

# Create admission datetime
df['admission_datetime'] = pd.to_datetime(df.apply(
    lambda row: f"{row['admission_date'].strftime('%Y-%m-%d')} {row['admission_time_only']}" if pd.notna(row['admission_date']) and pd.notna(row['admission_time_only']) else None, axis=1
), errors='coerce')

df['discharge_time'] = pd.to_datetime(df['discharge_time'], errors='coerce')

# Calculate LOS (Length of Stay) in days
df['length_of_stay'] = (df['discharge_time'] - df['admission_datetime']).dt.total_seconds() / (24 * 3600)

# Parse demographics
df['age'] = df['a/s'].str.extract(r'(\d+)').astype(float)
df['gender'] = df['a/s'].str.extract(r'/([MF])')

# Find ADD cases (Acute Diarrheal Disease)
add_keywords = ['gastroenteritis', 'gastro', 'diarrhea', 'diarrhoea', 'diarrh', 'dysentery', 'cholera', 'food poisoning', 'add', 'acute ge', 'age', 'diarrhrea', 'loose', 'motion', 'stool', 'bowel', 'enteric', 'dehydration']
add_cases = []
for idx, row in df.iterrows():
    diagnosis = str(row['diagnosis']).lower()
    if any(keyword in diagnosis for keyword in add_keywords):
        add_cases.append(idx)

add_df = df.loc[add_cases].copy() if add_cases else pd.DataFrame()

# Filter out invalid LOS values (negative or extremely high)
valid_add_los_df = add_df[(add_df['length_of_stay'] >= 0) & (add_df['length_of_stay'] <= 365)].copy()

print(f"ADD cases found: {len(add_df)}")
print(f"ADD cases with valid LOS: {len(valid_add_los_df)}")
print(f"ADD Mean LOS: {valid_add_los_df['length_of_stay'].mean():.1f} days")
print(f"ADD Median LOS: {valid_add_los_df['length_of_stay'].median():.1f} days")

# Create age groups
age_bins = [0, 5, 18, 35, 50, 65, 100]
age_labels = ['0-4', '5-17', '18-34', '35-49', '50-64', '65+']
valid_add_los_df['age_group'] = pd.cut(valid_add_los_df['age'], bins=age_bins, labels=age_labels, right=False)

# Create LOS categories
los_bins = [0, 1, 3, 7, 14, 30, 1000]
los_labels = ['1 day', '2-3 days', '4-7 days', '8-14 days', '15-30 days', '30+ days']
valid_add_los_df['los_category'] = pd.cut(valid_add_los_df['length_of_stay'], bins=los_bins, labels=los_labels, right=False)

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Sheet 1: Dashboard Summary
ws_summary = wb.create_sheet("Summary")

ws_summary['A1'] = "SIMSRH IPD - ADD Cases Length of Stay (LOS) Analysis"
ws_summary['A1'].font = Font(size=16, bold=True)
ws_summary.merge_cells('A1:D1')

# Key metrics
ws_summary['A3'] = "Key Metrics"
ws_summary['A3'].font = Font(size=14, bold=True)

metrics = [
    ["Total ADD Cases", len(add_df)],
    ["ADD Cases with Valid LOS", len(valid_add_los_df)],
    ["Mean Length of Stay", f"{valid_add_los_df['length_of_stay'].mean():.1f} days"],
    ["Median Length of Stay", f"{valid_add_los_df['length_of_stay'].median():.1f} days"],
    ["Min LOS", f"{valid_add_los_df['length_of_stay'].min():.1f} days"],
    ["Max LOS", f"{valid_add_los_df['length_of_stay'].max():.1f} days"],
    ["Date Range", "Aug 1 - Nov 12, 2025"],
]

for i, (metric, value) in enumerate(metrics, 5):
    ws_summary[f'A{i}'] = metric
    ws_summary[f'B{i}'] = value
    ws_summary[f'A{i}'].font = Font(bold=True)

# LOS Distribution
ws_summary['A15'] = "LOS Distribution by Category"
ws_summary['A15'].font = Font(size=14, bold=True)

los_dist = valid_add_los_df['los_category'].value_counts().sort_index()
for i, (category, count) in enumerate(los_dist.items(), 16):
    ws_summary[f'A{i}'] = str(category)
    ws_summary[f'B{i}'] = count
    ws_summary[f'C{i}'] = f"{count/len(valid_add_los_df)*100:.1f}%"

# Sheet 2: Raw Data with LOS
ws_data = wb.create_sheet("Raw Data")
data_cols = ['ip_number', 'diagnosis', 'department', 'age', 'gender', 'admission_datetime', 'discharge_time', 'length_of_stay', 'los_category']
for r, row in enumerate(dataframe_to_rows(valid_add_los_df[data_cols], index=False), 1):
    for c, value in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=value)

# Sheet 3: LOS Statistics
ws_stats = wb.create_sheet("LOS Statistics")

# LOS statistics
ws_stats['A1'] = "ADD Cases - Length of Stay Statistics"
ws_stats['A1'].font = Font(size=14, bold=True)

los_stats = [
    ["Mean LOS", valid_add_los_df['length_of_stay'].mean()],
    ["Median LOS", valid_add_los_df['length_of_stay'].median()],
    ["Std Deviation", valid_add_los_df['length_of_stay'].std()],
    ["Min LOS", valid_add_los_df['length_of_stay'].min()],
    ["25th Percentile", valid_add_los_df['length_of_stay'].quantile(0.25)],
    ["75th Percentile", valid_add_los_df['length_of_stay'].quantile(0.75)],
    ["Max LOS", valid_add_los_df['length_of_stay'].max()],
    ["Count", len(valid_add_los_df['length_of_stay'].dropna())],
]

for i, (stat, value) in enumerate(los_stats, 2):
    ws_stats[f'A{i}'] = stat
    ws_stats[f'B{i}'] = value

# LOS by Age Group
ws_stats['D1'] = "LOS by Age Group"
ws_stats['D1'].font = Font(size=14, bold=True)

age_los = valid_add_los_df.groupby('age_group')['length_of_stay'].agg(['mean', 'median', 'count']).round(1)
for i, (age_group, stats) in enumerate(age_los.iterrows(), 2):
    ws_stats[f'D{i}'] = str(age_group)
    ws_stats[f'E{i}'] = stats['mean']
    ws_stats[f'F{i}'] = stats['median']
    ws_stats[f'G{i}'] = stats['count']

# LOS by Gender
ws_stats['I1'] = "LOS by Gender"
ws_stats['I1'].font = Font(size=14, bold=True)

gender_los = valid_add_los_df.groupby('gender')['length_of_stay'].agg(['mean', 'median', 'count']).round(1)
for i, (gender, stats) in enumerate(gender_los.iterrows(), 2):
    ws_stats[f'I{i}'] = gender
    ws_stats[f'J{i}'] = stats['mean']
    ws_stats[f'K{i}'] = stats['median']
    ws_stats[f'L{i}'] = stats['count']

# LOS by Department
ws_stats['A15'] = "LOS by Department"
ws_stats['A15'].font = Font(size=14, bold=True)

dept_los = valid_add_los_df.groupby('department')['length_of_stay'].agg(['mean', 'median', 'count']).round(1).sort_values('mean', ascending=False)
for i, (dept, stats) in enumerate(dept_los.iterrows(), 16):
    ws_stats[f'A{i}'] = dept[:30] + "..." if len(str(dept)) > 30 else str(dept)
    ws_stats[f'B{i}'] = stats['mean']
    ws_stats[f'C{i}'] = stats['median']
    ws_stats[f'D{i}'] = stats['count']

# Sheet 4: Charts Data
ws_charts = wb.create_sheet("Charts Data")

# LOS distribution data
ws_charts['A1'] = "LOS Distribution by Category"
los_cat_data = valid_add_los_df['los_category'].value_counts().sort_index().reset_index()
los_cat_data.columns = ['LOS Category', 'Count']
for r, row in enumerate(dataframe_to_rows(los_cat_data, index=False), 2):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Age group LOS data
ws_charts['A15'] = "LOS by Age Group"
age_los_data = valid_add_los_df.groupby('age_group')['length_of_stay'].mean().reset_index()
age_los_data.columns = ['Age Group', 'Mean LOS']
for r, row in enumerate(dataframe_to_rows(age_los_data, index=False), 16):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Department LOS data (top 10)
ws_charts['A25'] = "LOS by Department (Top 10)"
dept_los_data = valid_add_los_df.groupby('department')['length_of_stay'].mean().sort_values(ascending=False).head(10).reset_index()
dept_los_data.columns = ['Department', 'Mean LOS']
for r, row in enumerate(dataframe_to_rows(dept_los_data, index=False), 26):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Save workbook
wb.save('ADD_LOS_Analysis_Dashboard.xlsx')
print("ADD LOS Dashboard Excel file created successfully!")

# Create charts
plt.style.use('default')
sns.set_palette("husl")

# ADD LOS distribution histogram
plt.figure(figsize=(12, 8))
plt.hist(valid_add_los_df['length_of_stay'], bins=20, edgecolor='black', alpha=0.7, color='orange')
plt.xlabel('Length of Stay (days)')
plt.ylabel('Number of ADD Patients')
plt.title('Distribution of Length of Stay - Acute Diarrheal Disease Cases', fontsize=14, fontweight='bold')
plt.grid(True, alpha=0.3)
plt.axvline(valid_add_los_df['length_of_stay'].mean(), color='red', linestyle='--', linewidth=2,
           label=f'Mean: {valid_add_los_df["length_of_stay"].mean():.1f} days')
plt.axvline(valid_add_los_df['length_of_stay'].median(), color='green', linestyle='--', linewidth=2,
           label=f'Median: {valid_add_los_df["length_of_stay"].median():.1f} days')
plt.legend()
plt.tight_layout()
plt.savefig('add_los_distribution.png', dpi=300, bbox_inches='tight')
plt.close()

# ADD LOS by age group
plt.figure(figsize=(10, 6))
age_los_means = valid_add_los_df.groupby('age_group')['length_of_stay'].mean().sort_index()
bars = plt.bar(range(len(age_los_means)), age_los_means.values, color='lightcoral', edgecolor='black', alpha=0.8)
plt.xticks(range(len(age_los_means)), age_los_means.index, rotation=45)
plt.xlabel('Age Group')
plt.ylabel('Average Length of Stay (days)')
plt.title('Average LOS by Age Group - ADD Cases')
plt.grid(True, alpha=0.3)

for bar, value in zip(bars, age_los_means.values):
    plt.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.1,
             f'{value:.1f}', ha='center', va='bottom', fontweight='bold')

plt.tight_layout()
plt.savefig('add_los_by_age_group.png', dpi=300, bbox_inches='tight')
plt.close()

# ADD LOS by gender
plt.figure(figsize=(8, 6))
gender_los_means = valid_add_los_df.groupby('gender')['length_of_stay'].mean()
bars = plt.bar(gender_los_means.index, gender_los_means.values, color=['lightblue', 'lightpink'], edgecolor='black', alpha=0.8)
plt.xlabel('Gender')
plt.ylabel('Average Length of Stay (days)')
plt.title('Average LOS by Gender - ADD Cases')
plt.grid(True, alpha=0.3)

for bar, value in zip(bars, gender_los_means.values):
    plt.text(bar.get_x(), bar.get_height() + 0.1,
             f'{value:.1f}', ha='center', va='bottom', fontweight='bold')

plt.tight_layout()
plt.savefig('add_los_by_gender.png', dpi=300, bbox_inches='tight')
plt.close()

# ADD LOS categories pie chart
plt.figure(figsize=(10, 8))
los_cat_counts = valid_add_los_df['los_category'].value_counts()
colors = ['#ff9999','#66b3ff','#99ff99','#ffcc99','#c2c2f0','#ffb3e6']
explode = [0.05] * len(los_cat_counts)

plt.pie(los_cat_counts.values, labels=los_cat_counts.index, autopct='%1.1f%%',
        colors=colors[:len(los_cat_counts)], explode=explode, shadow=True, startangle=90)
plt.title('ADD Cases - Length of Stay Distribution by Category', fontsize=14, fontweight='bold')
plt.axis('equal')
plt.tight_layout()
plt.savefig('add_los_categories_pie.png', dpi=300, bbox_inches='tight')
plt.close()

print("ADD LOS dashboard and charts created successfully!")
print("Files created:")
print("- ADD_LOS_Analysis_Dashboard.xlsx")
print("- add_los_distribution.png")
print("- add_los_by_age_group.png")
print("- add_los_by_gender.png")
print("- add_los_categories_pie.png")

# Print key findings for manuscript
print("\nKey ADD LOS Findings for Manuscript:")
print("="*50)
print(f"- ADD cases analyzed: {len(valid_add_los_df)}")
print(".1f")
print(".1f")
print(".1f")
print(".1f")
print("- Age group with longest LOS: ", end="")
max_age_group = valid_add_los_df.groupby('age_group')['length_of_stay'].mean().idxmax()
max_age_los = valid_add_los_df.groupby('age_group')['length_of_stay'].mean().max()
print(f"{max_age_group} ({max_age_los:.1f} days)")
print("- Gender with longer LOS: ", end="")
max_gender = valid_add_los_df.groupby('gender')['length_of_stay'].mean().idxmax()
max_gender_los = valid_add_los_df.groupby('gender')['length_of_stay'].mean().max()
print(f"{max_gender} ({max_gender_los:.1f} days)")

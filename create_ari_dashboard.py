import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

print("Creating ARI dashboard...")

# Load and prepare data
df = pd.read_excel('Compiled IPD case data SIMSRH_4months.xls')
df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

# Extract admission date
def extract_date_from_ip(ip_num):
    if pd.isna(ip_num):
        return None
    ip_str = str(ip_num)
    if len(ip_str) >= 9 and ip_str.startswith('IP'):
        try:
            year = 2000 + int(ip_str[2:4])
            month = int(ip_str[4:6])
            day = int(ip_str[6:8])
            return datetime(year, month, day)
        except:
            return None
    return None

df['admission_date'] = df['ip_number'].apply(extract_date_from_ip)

# Parse demographics
df['age'] = df['a/s'].str.extract(r'(\d+)').astype(float)
df['gender'] = df['a/s'].str.extract(r'/([MF])')

# Calculate LOS
df['admission_time_only'] = pd.to_datetime(df['admission_time'], errors='coerce').dt.time
df['admission_datetime'] = pd.to_datetime(df.apply(
    lambda row: f"{row['admission_date'].strftime('%Y-%m-%d')} {row['admission_time_only']}" if pd.notna(row['admission_date']) and pd.notna(row['admission_time_only']) else None, axis=1
), errors='coerce')
df['discharge_time'] = pd.to_datetime(df['discharge_time'], errors='coerce')
df['length_of_stay'] = (df['discharge_time'] - df['admission_datetime']).dt.total_seconds() / (24 * 3600)

# Find ARI cases (comprehensive search)
ari_keywords = ['ari', 'arti', 'urti', 'lrti', 'respiratory', 'viral fever', 'fever', 'bronchiolitis', 'pneumonia', 'cough', 'breath', 'lung', 'resp', 'acute febrile', 'febrile illness', 'bronchitis', 'pharyngitis', 'sinusitis', 'otitis', 'tonsillitis', 'viral fever', 'acute febrile illness', 'lower respiratory tract infection', 'upper respiratory tract infection']
ari_cases = []
for idx, row in df.iterrows():
    diagnosis = str(row['diagnosis']).lower()
    if any(keyword in diagnosis for keyword in ari_keywords):
        ari_cases.append(idx)

ari_df = df.loc[ari_cases].copy() if ari_cases else pd.DataFrame()
print(f"Found {len(ari_df)} ARI cases")

# Create age groups
if len(ari_df) > 0:
    age_bins = [0, 5, 18, 35, 50, 65, 100]
    age_labels = ['0-4', '5-17', '18-34', '35-49', '50-64', '65+']
    ari_df['age_group'] = pd.cut(ari_df['age'], bins=age_bins, labels=age_labels, right=False)

# Create Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Sheet 1: Dashboard Summary
ws_summary = wb.create_sheet("Summary")

ws_summary['A1'] = "SIMSRH IPD - Acute Respiratory Infection (ARI) Dashboard"
ws_summary['A1'].font = Font(size=16, bold=True)
ws_summary.merge_cells('A1:D1')

# Key metrics
ws_summary['A3'] = "Key Metrics"
ws_summary['A3'].font = Font(size=14, bold=True)

metrics = [
    ["Total ARI Cases", len(ari_df)],
    ["Percentage of Admissions", f"{len(ari_df)/len(df)*100:.1f}%"],
    ["Date Range", "Aug 1 - Nov 12, 2025"],
    ["Mean Age", f"{ari_df['age'].mean():.1f}"],
    ["Median Age", f"{ari_df['age'].median():.1f}"],
    ["Male Cases", (ari_df['gender'] == 'M').sum()],
    ["Female Cases", (ari_df['gender'] == 'F').sum()],
    ["Average LOS", f"{ari_df['length_of_stay'].mean():.1f}"],
]

for i, (metric, value) in enumerate(metrics, 5):
    ws_summary[f'A{i}'] = metric
    ws_summary[f'B{i}'] = value
    ws_summary[f'A{i}'].font = Font(bold=True)

# Sheet 2: Raw Data
ws_data = wb.create_sheet("Raw Data")
data_cols = ['ip_number', 'diagnosis', 'department', 'age', 'gender', 'admission_date', 'length_of_stay']
for r, row in enumerate(dataframe_to_rows(ari_df[data_cols], index=False), 1):
    for c, value in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=value)

# Sheet 3: Statistics
ws_stats = wb.create_sheet("Statistics")

# Age statistics
ws_stats['A1'] = "Age Statistics"
ws_stats['A1'].font = Font(size=14, bold=True)

age_stats = [
    ["Mean Age", ari_df['age'].mean()],
    ["Median Age", ari_df['age'].median()],
    ["Std Deviation", ari_df['age'].std()],
    ["Min Age", ari_df['age'].min()],
    ["Max Age", ari_df['age'].max()],
    ["Count", len(ari_df['age'].dropna())],
]

for i, (stat, value) in enumerate(age_stats, 2):
    ws_stats[f'A{i}'] = stat
    ws_stats[f'B{i}'] = value

# Gender distribution
ws_stats['D1'] = "Gender Distribution"
ws_stats['D1'].font = Font(size=14, bold=True)

gender_stats = [
    ["Male", (ari_df['gender'] == 'M').sum(), f"{(ari_df['gender'] == 'M').sum()/len(ari_df)*100:.1f}%"],
    ["Female", (ari_df['gender'] == 'F').sum(), f"{(ari_df['gender'] == 'F').sum()/len(ari_df)*100:.1f}%"],
    ["Total", len(ari_df), "100.0%"],
]

for i, (gender, count, pct) in enumerate(gender_stats, 2):
    ws_stats[f'D{i}'] = gender
    ws_stats[f'E{i}'] = count
    ws_stats[f'F{i}'] = pct

# Age group distribution
ws_stats['H1'] = "Age Group Distribution"
ws_stats['H1'].font = Font(size=14, bold=True)

if len(ari_df) > 0:
    age_group_stats = ari_df['age_group'].value_counts().sort_index()
    for i, (group, count) in enumerate(age_group_stats.items(), 2):
        ws_stats[f'H{i}'] = str(group)
        ws_stats[f'I{i}'] = count
        ws_stats[f'J{i}'] = f"{count/len(ari_df)*100:.1f}%"

# Department distribution
ws_stats['A15'] = "Department Distribution"
ws_stats['A15'].font = Font(size=14, bold=True)

dept_stats = ari_df['department'].value_counts()
for i, (dept, count) in enumerate(dept_stats.items(), 16):
    ws_stats[f'A{i}'] = dept
    ws_stats[f'B{i}'] = count
    ws_stats[f'C{i}'] = f"{count/len(ari_df)*100:.1f}%"

# Top diagnoses
ws_stats['E15'] = "Top 10 Diagnoses"
ws_stats['E15'].font = Font(size=14, bold=True)

top_dx = ari_df['diagnosis'].value_counts().head(10)
for i, (dx, count) in enumerate(top_dx.items(), 16):
    ws_stats[f'E{i}'] = dx[:50] + "..." if len(dx) > 50 else dx
    ws_stats[f'F{i}'] = count
    ws_stats[f'G{i}'] = f"{count/len(ari_df)*100:.1f}%"

# Monthly trends
if len(ari_df) > 0:
    ari_df['admission_month'] = ari_df['admission_datetime'].dt.to_period('M')
    monthly_stats = ari_df.groupby('admission_month').size()

    ws_stats['A25'] = "Monthly Trends"
    ws_stats['A25'].font = Font(size=14, bold=True)

    for i, (month, count) in enumerate(monthly_stats.items(), 26):
        ws_stats[f'A{i}'] = str(month)
        ws_stats[f'B{i}'] = count

# Sheet 4: Charts Data
ws_charts = wb.create_sheet("Charts Data")

# Age group data
ws_charts['A1'] = "Age Group Distribution"
if len(ari_df) > 0:
    age_group_data = ari_df['age_group'].value_counts().sort_index().reset_index()
    age_group_data.columns = ['Age Group', 'Count']
    for r, row in enumerate(dataframe_to_rows(age_group_data, index=False), 2):
        for c, value in enumerate(row, 1):
            ws_charts.cell(row=r, column=c, value=value)

# Department data
ws_charts['A10'] = "Department Distribution"
dept_data = ari_df['department'].value_counts().reset_index()
dept_data.columns = ['Department', 'Count']
for r, row in enumerate(dataframe_to_rows(dept_data, index=False), 11):
    for c, value in enumerate(row, 1):
        ws_charts.cell(row=r, column=c, value=value)

# Monthly data
if len(ari_df) > 0:
    ws_charts['A20'] = "Monthly Distribution"
    monthly_data = ari_df.groupby('admission_month').size().reset_index()
    monthly_data.columns = ['Month', 'Cases']
    monthly_data['Month'] = monthly_data['Month'].astype(str)
    for r, row in enumerate(dataframe_to_rows(monthly_data, index=False), 21):
        for c, value in enumerate(row, 1):
            ws_charts.cell(row=r, column=c, value=value)

# Save workbook
wb.save('ARI_Dashboard.xlsx')
print("ARI Dashboard Excel file created successfully!")

# Create charts
plt.style.use('default')
sns.set_palette("husl")

# Age distribution
plt.figure(figsize=(10, 6))
plt.hist(ari_df['age'].dropna(), bins=15, edgecolor='black', alpha=0.7, color='lightcoral')
plt.xlabel('Age (years)')
plt.ylabel('Number of ARI Cases')
plt.title('Age Distribution of Acute Respiratory Infection Cases')
plt.grid(True, alpha=0.3)
plt.axvline(ari_df['age'].mean(), color='red', linestyle='--', linewidth=2,
           label=f'Mean: {ari_df["age"].mean():.1f} years')
plt.legend()
plt.tight_layout()
plt.savefig('ari_age_distribution.png', dpi=300, bbox_inches='tight')
plt.close()

# Gender pie chart
plt.figure(figsize=(8, 8))
gender_counts = ari_df['gender'].value_counts()
colors = ['lightblue', 'lightpink']
explode = (0.05, 0)

plt.pie(gender_counts.values, labels=gender_counts.index, autopct='%1.1f%%',
        colors=colors, explode=explode, shadow=True, startangle=90)
plt.title('Gender Distribution in ARI Cases', fontsize=14, fontweight='bold')
plt.axis('equal')
plt.tight_layout()
plt.savefig('ari_gender_pie.png', dpi=300, bbox_inches='tight')
plt.close()

# Age group bar chart
if len(ari_df) > 0:
    plt.figure(figsize=(10, 6))
    age_group_counts = ari_df['age_group'].value_counts().sort_index()
    bars = plt.bar(range(len(age_group_counts)), age_group_counts.values, color='teal', edgecolor='black', alpha=0.8)
    plt.xticks(range(len(age_group_counts)), age_group_counts.index, rotation=45)
    plt.xlabel('Age Group')
    plt.ylabel('Number of Cases')
    plt.title('ARI Cases by Age Group')
    plt.grid(True, alpha=0.3)

    for bar, count in zip(bars, age_group_counts.values):
        plt.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 2,
                 f'{int(count)}', ha='center', va='bottom', fontweight='bold')

    plt.tight_layout()
    plt.savefig('ari_age_groups.png', dpi=300, bbox_inches='tight')
    plt.close()

# Monthly trends
if len(ari_df) > 0:
    plt.figure(figsize=(12, 6))
    monthly_counts = ari_df.groupby('admission_month').size()
    plt.plot(range(len(monthly_counts)), monthly_counts.values, marker='o', linewidth=3,
             markersize=10, color='darkred', markerfacecolor='red', markeredgecolor='darkred')
    plt.xticks(range(len(monthly_counts)), [str(x) for x in monthly_counts.index], rotation=45)
    plt.xlabel('Month')
    plt.ylabel('Number of ARI Cases')
    plt.title('Monthly Trends of ARI Cases')
    plt.grid(True, alpha=0.3)

    for i, v in enumerate(monthly_counts.values):
        plt.text(i, v + 1, str(v), ha='center', va='bottom', fontweight='bold', fontsize=12)

    plt.tight_layout()
    plt.savefig('ari_monthly_trends.png', dpi=300, bbox_inches='tight')
    plt.close()

print("ARI dashboard and charts created successfully!")
print("Files created:")
print("- ARI_Dashboard.xlsx")
print("- ari_age_distribution.png")
print("- ari_gender_pie.png")
print("- ari_age_groups.png")
print("- ari_monthly_trends.png")
